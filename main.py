import openpyxl
import csv
from pathlib import Path
from Person import Person





############################################
#Beginning data parsing
xlsx_file = Path('school_record.xlsx')

wb_obj = openpyxl.load_workbook(xlsx_file)


#List of club names
eC = [
"Board Game Club",
"Football",
"Soccer",
"Video Game Club",
"Band",
"Computer Science Club",
"Choir",
"Basketball",
"Badminton",
"Baseball",
"Drama Club"
]

afterSchool ={}
for e in eC:
  afterSchool[e] = []


#List of classes
classes = [
    "Physics A",
    "Physics B",
    "Biology A",
    "Biology B",
    "Functions A",
    "Functions B",
    "Calculus A",
    "Calculus B",
    "Philosophy A",
    "Philosophy B",
    "Art A",
    "Art B",
    "Drama A",
    "Drama B",
    "Computer Science A",
    "Computer Science B",
    "Computer Engineering A",
    "Computer Engineering B",
    "Humanities A",
    "Humanities B",
]

periodOne = {}
periodTwo = {}
periodThree = {}
periodFour = {}

#adding class names to dict
for c in classes:
    periodOne[c] = []
    periodTwo[c] = []
    periodThree[c] = []
    periodFour[c] = []

worksheet = wb_obj.worksheets[0]
students = []
grade9 = []
grade10 = []
grade11 = []
grade12 = []

#Adding students to student list
for row in worksheet.iter_rows(max_row=581):
    if (isinstance(row[0].value, float)):
        newPerson = Person(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value,
                           row[5].value, row[6].value, row[7].value, row[8].value, row[9].value)
        students.append(newPerson)
        periodOne[newPerson.periodOne].append(newPerson)
        periodTwo[newPerson.periodTwo].append(newPerson)
        periodThree[newPerson.periodThree].append(newPerson)
        periodFour[newPerson.periodFour].append(newPerson)

        if newPerson.grade == 9:
          grade9.append(newPerson)
        if newPerson.grade == 10:
          grade10.append(newPerson)
        if newPerson.grade == 11:
          grade11.append(newPerson)
        if newPerson.grade == 12:
          grade12.append(newPerson)

        if newPerson.activities != "N/A":
          afterSchool[newPerson.activities.split(",")[0]].append(newPerson)

#adding teachers to teacher list
worksheet = wb_obj.worksheets[1]
teachers = []
for row in worksheet.iter_rows(max_row=21):
    if (isinstance(row[0].value, float)):
        newPerson = Person(row[0].value, row[1].value, row[2].value, 0, row[3].value,
                           row[3].value, row[3].value, row[3].value, "N/A", "N/A")
        teachers.append(newPerson)

        periodOne[newPerson.periodOne].append(newPerson)
        periodTwo[newPerson.periodTwo].append(newPerson)
        periodThree[newPerson.periodThree].append(newPerson)
        periodFour[newPerson.periodFour].append(newPerson)

worksheet = wb_obj.worksheets[2]
assistants = []
id_ctr = 1
for row in worksheet.iter_rows(max_row=7):
    if (row[0].value != "Last Name"):
        newPerson = Person(id_ctr, row[0].value, row[1].value, 0, row[2].value,
                           row[3].value, row[4].value, row[5].value, "N/A", "N/A")
        id_ctr += 1
        assistants.append(newPerson)

        periodOne[newPerson.periodOne].append(newPerson)
        periodTwo[newPerson.periodTwo].append(newPerson)
        periodThree[newPerson.periodThree].append(newPerson)
        periodFour[newPerson.periodFour].append(newPerson)

worksheet = wb_obj.worksheets[3]
infectedList = []





## Adding infected students
students[530].infected = True
students[530].probability =1.0
infectedList.append(students[530])
students[85].infected = True
students[85].probability =1.0
infectedList.append(students[85])
students[130].infected = True
students[130].probability =1.0
infectedList.append(students[130])
assistants[4].infected = True
assistants[4].probability =1.0
infectedList.append(assistants[4])
######################################################################
#Data parsing and insertion done



#Finding infection based on relation (Before School)
temp = []

for i in infectedList:
    for s in students:
      if i.lastName == s.lastName and s not in infectedList:
        s.infected = True
        s.probability =1.0
        temp.append(s)

infectedList += temp

#Find new infected
def newInfected(peopleInClass, infectedInClass):
  tempPerson = Person()
  largest =0
  for i in infectedInClass:
    if i.probability >largest:
      tempPerson = i
      largest = i.probability
  
  #Finding 3 most likely people to be infected
  highestWeights = [Person(), Person(), Person()]
  for i in range(3):
    minP =0
    for p in peopleInClass:
      if p.probability>minP and p not in highestWeights and p not in infectedList:
        minP = p.probability
        highestWeights[i]= p
        for j in infectedInClass:                                                       #
         if j.probability>0.85 and j.firstName +" "+ j.lastName not in p.infectedBy:    #
          p.infectedBy.append(j.firstName +" "+ j.lastName)                             #Finding who infection was most likely to be
        if len(p.infectedBy)==0:                                                        #caused by
          p.infectedBy.append(tempPerson.firstName + " " +tempPerson.lastName)          #
  for h in highestWeights:
    h.infected = True
    if h.probability !=0:
      infectedList.append(h)



  #Helper method to calculate probability
def updateProbabilityForClass(peopleInClass):
  prob = 3/len(peopleInClass)
  t = []
  for p in peopleInClass:
    if not p.infected:
      if p.multiplier != 0: 
        p.probability += prob*p.multiplier  #Calculating probability
      else:
        p.probability += prob
      if p.probability >=1:
        p.probability =1.0
    else:
      t.append(p) #returning a list of the students in this class
  return t


#Finding classes with infected students and updating probability
def updateProbability(period, infectedList):
  classList = []
  #Creating list of classes with infected students
  for infectedPerson in infectedList:
    if (period == 1):
      classList.append(infectedPerson.periodOne)
    elif (period == 2):
      classList.append(infectedPerson.periodTwo)
    elif (period == 3):
      classList.append(infectedPerson.periodThree)
    elif (period == 4):
      classList.append(infectedPerson.periodFour)
    elif(period==5):
      classList.append(infectedPerson.activities)
    else:
      continue
  
  #Selecting which period to update probability for
  if (period == 1):
    for c in classList:
      infectedInClass = updateProbabilityForClass(periodOne[c])
      newInfected(periodOne[c], infectedInClass)
  elif (period == 2):
    for c in classList:
      infectedInClass = updateProbabilityForClass(periodTwo[c])
      newInfected(periodTwo[c], infectedInClass)
  elif (period == 3):
    for c in classList:
      infectedInClass = updateProbabilityForClass(periodThree[c])
      newInfected(periodThree[c], infectedInClass)
  elif (period == 4):
    for c in classList:
      if c in periodFour.keys():
        infectedInClass = updateProbabilityForClass(periodFour[c])
        newInfected(periodFour[c], infectedInClass)  
  elif (period == 5):
    for c in classList:
      if c in afterSchool.keys():
        infectedInClass = updateProbabilityForClass(afterSchool[c])
        newInfected(afterSchool[c], infectedInClass)
  elif (period == 9):
    infectedInClass = updateProbabilityForClass(grade9)
    newInfected(grade9, infectedInClass)
  elif (period == 10):
    infectedInClass = updateProbabilityForClass(grade10)
    newInfected(grade10, infectedInClass)
  elif (period == 11):
    infectedInClass = updateProbabilityForClass(grade11)
    newInfected(grade11, infectedInClass)
  elif (period == 12):
    infectedInClass = updateProbabilityForClass(grade12)
    newInfected(grade12, infectedInClass)



##############################################
#Calling functions for all periods, lunch, and after school activities
updateProbability(1, infectedList)
updateProbability(2, infectedList)
updateProbability(3, infectedList)
updateProbability(4, infectedList)
updateProbability(5, infectedList)
updateProbability(9, infectedList)
updateProbability(10, infectedList)
updateProbability(11, infectedList)
updateProbability(12, infectedList)



#Finding infection based on relation (After School)
temp = []

for i in infectedList:
    for s in students:
      if i.lastName == s.lastName and s not in infectedList:
        s.infected = True
        s.probability =i.probability
        temp.append(s)

infectedList += temp

print("--------------------")
print("      Teachers")
print("--------------------")
for t in teachers:
  t.probability = round(t.probability, 3)
  print(t.probability, t.firstName, t.lastName, "Infected by:", t.infectedBy)
print("------------------------------")
print("      Teaching Assistants")
print("------------------------------")
for t in assistants:
  t.probability = round(t.probability, 3)
  print(t.probability, t.firstName, t.lastName, "Infected by:", t.infectedBy)
print("--------------------")
print("      Students")
print("--------------------")
for i in students:
  i.probability = round(i.probability, 3)
  print(i.probability, i.firstName, i.lastName, "Infected by:", i.infectedBy)

